import requests
import pandas as pd
import numpy as np
from datetime import datetime, timezone as dt_timezone
from pytz import timezone
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt

# === USER CONFIG === #
OANDA_TOKEN = "37ee33b35f88e073a08d533849f7a24b-524c89ef15f36cfe532f0918a6aee4c2"
OANDA_ACCOUNT_ID = "101-004-35770497-001"
SYMBOL = "XAU_USD"
GRANULARITY = "M5"
CANDLES = 3840
LOT_SIZE = 1.0
START_BALANCE = 100000.0
BACKTEST_START_DATE = "2025-01-01T00:00:00Z"  # <-- Use ISO 8601 format
LOCAL_TZ = timezone('Europe/London')
EXPORT_DIR = r"C:\Users\anish\OneDrive\Desktop\Anish\Trading Bots\EMA-VWAP Scalp"

# === FETCH DATA === #
def fetch_oanda_candles():
    max_per_request = 5000
    oanda_url = "https://api-fxpractice.oanda.com/v3/instruments"
    headers = {"Authorization": f"Bearer {OANDA_TOKEN}"}
    all_data = []

    start_time = pd.to_datetime(BACKTEST_START_DATE)
    end_time = datetime.now(dt_timezone.utc).replace(microsecond=0)

    while start_time < end_time:
        params = {
            "granularity": GRANULARITY,
            "price": "M",
            "from": start_time.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "count": max_per_request
        }

        url = f"{oanda_url}/{SYMBOL}/candles"
        print(f"[*] Fetching candles starting {start_time}")

        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            raise RuntimeError(f"Failed to fetch candles: {response.status_code} - {response.text}")

        candles = response.json().get("candles", [])
        if not candles:
            break

        for c in candles:
            if c['complete']:
                utc_time = pd.to_datetime(c['time']).tz_convert(LOCAL_TZ)
                all_data.append({
                    "time": utc_time,
                    "open": float(c['mid']['o']),
                    "high": float(c['mid']['h']),
                    "low": float(c['mid']['l']),
                    "close": float(c['mid']['c']),
                    "volume": int(c['volume'])
                })

        # Move start_time to last candle fetched + 1 second
        last_time = pd.to_datetime(candles[-1]['time'])
        start_time = last_time + pd.Timedelta(seconds=1)

    if not all_data:
        raise RuntimeError("No candle data retrieved from OANDA.")

    df = pd.DataFrame(all_data)
    df.set_index("time", inplace=True)
    return df

# === CALCULATE INDICATORS === #
def calculate_indicators(df):
    df['ema20'] = df['close'].ewm(span=20, adjust=False).mean()
    df['ema50'] = df['close'].ewm(span=50, adjust=False).mean()
    df['vwap'] = (df['volume'] * (df['high']+df['low']+df['close'])/3).cumsum() / df['volume'].cumsum()
    df['tr'] = np.maximum(df['high']-df['low'], np.maximum(abs(df['high']-df['close'].shift(1)), abs(df['low']-df['close'].shift(1))))
    df['atr'] = df['tr'].rolling(window=14).mean()
    return df

# === BACKTEST STRATEGY === #
def run_backtest(df):
    trades = []
    balance = START_BALANCE
    position = None
    entry_price = 0
    sl = 0
    tp = 0

    for i in range(51, len(df)):
        price = df['close'].iloc[i]
        row = df.iloc[i]

        # ENTRY CONDITIONS
        if position is None:
            # LONG SETUP
            if row['close'] > row['vwap'] and row['ema20'] > row['ema50'] and row['close'] > row['ema20'] and df['low'].iloc[i-1] <= row['ema20']:
                risk = row['atr']
                sl = price - risk
                tp = price + 1.5 * (price - sl)
                position = 'BUY'
                entry_price = price
                entry_time = df.index[i]

            # SHORT SETUP
            elif row['close'] < row['vwap'] and row['ema20'] < row['ema50'] and row['close'] < row['ema20'] and df['high'].iloc[i-1] >= row['ema20']:
                risk = row['atr']
                sl = price + risk
                tp = price - 1.5 * (sl - price)
                position = 'SELL'
                entry_price = price
                entry_time = df.index[i]

        # EXIT CONDITIONS
        else:
            if position == 'BUY' and (row['low'] <= sl or row['high'] >= tp):
                exit_price = sl if row['low'] <= sl else tp
                profit = (exit_price - entry_price) * LOT_SIZE
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': 'BUY', 'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

            elif position == 'SELL' and (row['high'] >= sl or row['low'] <= tp):
                exit_price = sl if row['high'] >= sl else tp
                profit = (entry_price - exit_price) * LOT_SIZE
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': 'SELL', 'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

    trades_df = pd.DataFrame(trades)
    return trades_df, balance

# === SUMMARIZE RESULTS === #
def summarize_results(trades, final_balance):
    ticker = SYMBOL
    timeframe = GRANULARITY
    start_time = trades['entry_time'].min()
    end_time = trades['exit_time'].max()

    trades_fmt = trades.copy()
    trades_fmt['entry_time'] = trades_fmt['entry_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['exit_time'] = trades_fmt['exit_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['entry_price'] = trades_fmt['entry_price'].round(2)
    trades_fmt['exit_price'] = trades_fmt['exit_price'].round(2)
    trades_fmt['profit'] = trades_fmt['profit'].round(2)

    trades['equity'] = START_BALANCE + trades['profit'].cumsum()
    plt.figure(figsize=(10, 4))
    plt.plot(trades['exit_time'], trades['equity'], label="Equity", linewidth=2)
    plt.title(f"Equity Curve - {ticker} ({timeframe})")
    plt.xlabel("Time")
    plt.ylabel("Account Equity ($)")
    plt.grid(True)
    plt.tight_layout()
    curve_path = os.path.join(EXPORT_DIR, f"equity_curve_{ticker}_{timeframe}.png")
    plt.savefig(curve_path)
    plt.close()

    summary = {
        'Ticker': ticker,
        'Timeframe': timeframe,
        'Start Time': start_time.strftime("%d/%m/%Y %H:%M:%S"),
        'End Time': end_time.strftime("%d/%m/%Y %H:%M:%S"),
        'Starting Balance': START_BALANCE,
        'Ending Balance': round(final_balance, 2),
        'Total Return': round(final_balance - START_BALANCE, 2),
        'Total Trades': len(trades),
        'Biggest Win': round(trades['profit'].max(), 2),
        'Biggest Loss': round(trades['profit'].min(), 2),
        'Winning Trades': int((trades['profit'] > 0).sum()),
        'Win Rate (%)': round((trades['profit'] > 0).mean(), 4)
    }

    print("\n===== BACKTEST SUMMARY =====")
    for key, value in summary.items():
        print(f"{key:<20}: {value}")
    print("============================\n")

    os.makedirs(EXPORT_DIR, exist_ok=True)
    output_path = os.path.join(EXPORT_DIR, f"EMA_VWAP_Scalping_{ticker}_{timeframe}.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Trades"
    for r in dataframe_to_rows(trades_fmt, index=False, header=True):
        ws1.append(r)
    bold_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = bold_font
    ws2 = wb.create_sheet("Summary")
    for k, v in summary.items():
        ws2.append([k, v])
    ws3 = wb.create_sheet("Equity Curve")
    img = ExcelImage(curve_path)
    img.anchor = "A1"
    ws3.add_image(img)
    wb.save(output_path)
    print(f"\n[OK] Backtest saved to Excel: {output_path}")
    print(f"[OK] Equity curve image saved: {curve_path}")

if __name__ == "__main__":
    print("[*] Fetching OANDA data...")
    df = fetch_oanda_candles()
    df = calculate_indicators(df)
    print("[*] Running backtest...")
    trades, end_balance = run_backtest(df)
    summarize_results(trades, end_balance)
