import requests
import pandas as pd
import numpy as np
from datetime import datetime, timezone as dt_timezone
from pytz import timezone
import os
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt

# === USER CONFIG === #
OANDA_TOKEN = "37ee33b35f88e073a08d533849f7a24b-524c89ef15f36cfe532f0918a6aee4c2"
OANDA_ACCOUNT_ID = "101-004-35770497-001"
SYMBOL = "XAU_USD"
GRANULARITY = "M5"
CANDLES = 3840
START_BALANCE = 100000.0
BACKTEST_START_DATE = "2025-01-01T00:00:00Z"
LOCAL_TZ = timezone('Europe/London')
EXPORT_DIR = r"C:\\Users\\anish\\OneDrive\\Desktop\\Anish\\Trading Bots\\EMA-VWAP Parameter Test"

# === PARAMETER RANGES === #
EMA_FAST_VALUES = [9, 10]
EMA_SLOW_VALUES = [21, 50]
SL_MULT_VALUES = [1.5, 2.0]
TP_MULT_VALUES = [2.0, 2.5]
RISK_PER_TRADE_VALUES = [0.005, 0.01]


# === FETCH DATA === #
def fetch_oanda_candles():
    max_per_request = 5000
    oanda_url = "https://api-fxpractice.oanda.com/v3/instruments"
    headers = {"Authorization": f"Bearer {OANDA_TOKEN}"}
    all_data = []

    start_time = pd.to_datetime(BACKTEST_START_DATE)
    end_time = datetime.now(dt_timezone.utc).replace(microsecond=0)

    while start_time < end_time:
        params = {
            "granularity": GRANULARITY,
            "price": "M",
            "from": start_time.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "count": max_per_request
        }

        url = f"{oanda_url}/{SYMBOL}/candles"
        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            raise RuntimeError(f"Failed to fetch candles: {response.status_code} - {response.text}")

        candles = response.json().get("candles", [])
        if not candles:
            break

        for c in candles:
            if c['complete']:
                utc_time = pd.to_datetime(c['time']).tz_convert(LOCAL_TZ)
                all_data.append({
                    "time": utc_time,
                    "open": float(c['mid']['o']),
                    "high": float(c['mid']['h']),
                    "low": float(c['mid']['l']),
                    "close": float(c['mid']['c']),
                    "volume": int(c['volume'])
                })

        last_time = pd.to_datetime(candles[-1]['time'])
        start_time = last_time + pd.Timedelta(seconds=1)

    df = pd.DataFrame(all_data)
    df.set_index("time", inplace=True)
    return df


# === INDICATORS === #
def calculate_indicators(df, ema_fast, ema_slow):
    df['ema_fast'] = df['close'].ewm(span=ema_fast, adjust=False).mean()
    df['ema_slow'] = df['close'].ewm(span=ema_slow, adjust=False).mean()
    df['vwap'] = (df['volume'] * (df['high'] + df['low'] + df['close']) / 3).cumsum() / df['volume'].cumsum()
    df['tr'] = np.maximum(df['high'] - df['low'],
                          np.maximum(abs(df['high'] - df['close'].shift(1)), abs(df['low'] - df['close'].shift(1))))
    df['atr'] = df['tr'].rolling(window=14).mean()
    return df


# === BACKTEST === #
def run_backtest(df, ema_fast, ema_slow, sl_mult, tp_mult, risk_per_trade):
    df = calculate_indicators(df.copy(), ema_fast, ema_slow)
    trades = []
    balance = START_BALANCE
    position = None

    for i in range(51, len(df)):
        price = df['close'].iloc[i]
        row = df.iloc[i]

        if position is None:
            if (row['close'] > row['vwap'] and row['ema_fast'] > row['ema_slow'] and row['close'] > row['ema_fast'] and
                    df['low'].iloc[i - 1] <= row['ema_fast']):
                sl = price - sl_mult * row['atr']
                tp = price + tp_mult * (price - sl)
                lot_size = (balance * risk_per_trade) / (price - sl) if (price - sl) > 0 else 1.0
                position = ('BUY', price, sl, tp, df.index[i], lot_size)

            elif (row['close'] < row['vwap'] and row['ema_fast'] < row['ema_slow'] and row['close'] < row[
                'ema_fast'] and df['high'].iloc[i - 1] >= row['ema_fast']):
                sl = price + sl_mult * row['atr']
                tp = price - tp_mult * (sl - price)
                lot_size = (balance * risk_per_trade) / (sl - price) if (sl - price) > 0 else 1.0
                position = ('SELL', price, sl, tp, df.index[i], lot_size)

        else:
            direction, entry_price, sl, tp, entry_time, lot_size = position
            if direction == 'BUY' and (row['low'] <= sl or row['high'] >= tp):
                exit_price = sl if row['low'] <= sl else tp
                profit = (exit_price - entry_price) * lot_size
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': direction,
                               'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

            elif direction == 'SELL' and (row['high'] >= sl or row['low'] <= tp):
                exit_price = sl if row['high'] >= sl else tp
                profit = (entry_price - exit_price) * lot_size
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': direction,
                               'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

    return pd.DataFrame(trades), balance


# === GRID SEARCH TESTING === #
def run_grid_search(df):
    best_result = None
    best_config = None
    os.makedirs(EXPORT_DIR, exist_ok=True)

    for ema_fast, ema_slow, sl_mult, tp_mult, risk in itertools.product(EMA_FAST_VALUES, EMA_SLOW_VALUES,
                                                                        SL_MULT_VALUES, TP_MULT_VALUES,
                                                                        RISK_PER_TRADE_VALUES):
        trades, final_balance = run_backtest(df, ema_fast, ema_slow, sl_mult, tp_mult, risk)
        total_return = final_balance - START_BALANCE
        print(
            f"Tested: EMA({ema_fast}/{ema_slow}), SLx{sl_mult}, TPx{tp_mult}, Risk={risk * 100:.1f}% -> Return: ${total_return:,.2f}")

        if best_result is None or final_balance > best_result:
            best_result = final_balance
            best_config = (ema_fast, ema_slow, sl_mult, tp_mult, risk, trades)

    print("\n=== BEST CONFIGURATION FOUND ===")
    print(
        f"EMA: {best_config[0]}/{best_config[1]} | SLx: {best_config[2]} | TPx: {best_config[3]} | Risk: {best_config[4] * 100:.1f}%")
    summarize_results(best_config[5], best_result, best_config[:5])


# === REPORTING === #
def summarize_results(trades, final_balance, params):
    trades['equity'] = START_BALANCE + trades['profit'].cumsum()
    plt.figure(figsize=(10, 4))
    plt.plot(trades['exit_time'], trades['equity'], label="Equity", linewidth=2)
    plt.title("Equity Curve")
    plt.xlabel("Time")
    plt.ylabel("Account Equity ($)")
    plt.grid(True)
    plt.tight_layout()
    curve_path = os.path.join(EXPORT_DIR, f"best_equity_curve.png")
    plt.savefig(curve_path)
    plt.close()

    trades_fmt = trades.copy()
    trades_fmt['entry_time'] = trades_fmt['entry_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['exit_time'] = trades_fmt['exit_time'].dt.strftime("%d/%m/%Y %H:%M:%S")

    summary = {
        'EMA Fast': params[0],
        'EMA Slow': params[1],
        'SL Mult': params[2],
        'TP Mult': params[3],
        'Risk Per Trade': f"{params[4] * 100:.2f}%",
        'Final Balance': f"${final_balance:,.2f}",
        'Total Return': f"${final_balance - START_BALANCE:,.2f}",
        'Total Trades': len(trades_fmt),
        'Biggest Win': f"${trades['profit'].max():,.2f}",
        'Biggest Loss': f"${trades['profit'].min():,.2f}",
        'Win Rate (%)': f"{(trades['profit'] > 0).mean() * 100:.2f}%"
    }

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Trades"
    for r in dataframe_to_rows(trades_fmt, index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    ws2 = wb.create_sheet("Summary")
    for k, v in summary.items():
        ws2.append([k, v])
    ws3 = wb.create_sheet("Equity Curve")
    img = ExcelImage(curve_path)
    img.anchor = "A1"
    ws3.add_image(img)
    output_path = os.path.join(EXPORT_DIR, "Best_Parameter_Result.xlsx")
    wb.save(output_path)
    print(f"[OK] Best configuration report saved: {output_path}")


# === MAIN === #
if __name__ == "__main__":
    print("[*] Fetching OANDA data...")
    df = fetch_oanda_candles()
    print("[*] Running grid search across parameters...")
    run_grid_search(df)
