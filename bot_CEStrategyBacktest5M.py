import requests
import pandas as pd
from datetime import datetime, timezone as dt_timezone
from pytz import timezone
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, numbers
import matplotlib.pyplot as plt
import numpy as np


"""
Changes required for different timeframes:
GRANULARITY = "M5"      >> M5, M30, H1 etc...
CANDLES = 11520         >> For 2 months of 30-minute candles: Each day has 48 candles (24h × 2) 
                                                              ~40 trading days → 48 × 40 = ~1920 candles
"""


# === USER CONFIG === #
OANDA_TOKEN = "37ee33b35f88e073a08d533849f7a24b-524c89ef15f36cfe532f0918a6aee4c2"
OANDA_ACCOUNT_ID = "101-004-35770497-001"
SYMBOL = "XAU_USD"
GRANULARITY = "M15"
CANDLES = 3840     # 1H = 4 candles, (24H * 4 = 96), (5D * 96 = 480), (8W * 480 = 3840 candles for 2 months)
ATR_PERIOD = 1
ATR_MULT = 1.85
LOT_SIZE = 1.0
START_BALANCE = 100000.0
LOCAL_TZ = timezone('Europe/London')

# === CUSTOM OUTPUT DIRECTORY ===
EXPORT_DIR = r"C:\Users\anish\Desktop\Anish\CE Stretegy Backtest"  # <-- Change this to your target folder


def fetch_oanda_candles():
    max_per_request = 5000
    total_needed = 11520
    oanda_url = "https://api-fxpractice.oanda.com/v3/instruments"
    headers = {"Authorization": f"Bearer {OANDA_TOKEN}"}
    all_data = []

    end_time = datetime.now(dt_timezone.utc).replace(microsecond=0)
    collected = 0

    while collected < total_needed:
        params = {
            "granularity": GRANULARITY,
            "count": min(max_per_request, total_needed - collected),
            "price": "M",
            "to": end_time.strftime("%Y-%m-%dT%H:%M:%SZ")
        }

        url = f"{oanda_url}/{SYMBOL}/candles"
        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            raise RuntimeError(f"Failed to fetch candles: {response.status_code} - {response.text}")

        candles = response.json().get("candles", [])
        if not candles:
            break

        batch = []
        for c in candles:
            if c['complete']:
                utc_time = pd.to_datetime(c['time']).tz_convert(LOCAL_TZ)
                batch.append({
                    "time": utc_time,
                    "open": float(c['mid']['o']),
                    "high": float(c['mid']['h']),
                    "low": float(c['mid']['l']),
                    "close": float(c['mid']['c']),
                    "volume": int(c['volume'])
                })

        if not batch:
            break

        all_data = batch + all_data  # prepend to keep order
        end_time = pd.to_datetime(candles[0]['time'])  # move "to" back for next batch
        collected += len(batch)

    df = pd.DataFrame(all_data)
    df.set_index("time", inplace=True)
    return df


def calculate_heikin_ashi(df):
    ha = pd.DataFrame(index=df.index)
    ha['ha_close'] = (df['open'] + df['high'] + df['low'] + df['close']) / 4
    ha_open = [(df['open'].iloc[0] + df['close'].iloc[0]) / 2]
    for i in range(1, len(df)):
        ha_open.append((ha_open[i-1] + ha['ha_close'].iloc[i-1]) / 2)
    ha['ha_open'] = ha_open
    ha['ha_high'] = pd.concat([df['high'], ha['ha_open'], ha['ha_close']], axis=1).max(axis=1)
    ha['ha_low'] = pd.concat([df['low'], ha['ha_open'], ha['ha_close']], axis=1).min(axis=1)
    return ha


def calculate_indicators(df):
    ha = calculate_heikin_ashi(df)
    tr = pd.DataFrame(index=df.index)
    tr['ha_close'] = ha['ha_close']
    tr['ha_high'] = ha['ha_high']
    tr['ha_low'] = ha['ha_low']
    tr['ha_open'] = ha['ha_open']
    tr['prev_close'] = ha['ha_close'].shift(1)

    tr['true_range'] = tr.apply(lambda row: max(
        row['ha_high'] - row['ha_low'],
        abs(row['ha_high'] - row['prev_close']),
        abs(row['ha_low'] - row['prev_close'])
    ), axis=1)

    tr['atr'] = tr['true_range'].ewm(alpha=1/ATR_PERIOD, adjust=False).mean()

    long_stop = tr['ha_high'].rolling(window=ATR_PERIOD).max() - tr['atr'] * ATR_MULT
    short_stop = tr['ha_low'].rolling(window=ATR_PERIOD).min() + tr['atr'] * ATR_MULT

    long_smooth = long_stop.copy()
    short_smooth = short_stop.copy()

    for i in range(1, len(tr)):
        if tr['ha_close'].iloc[i-1] > long_smooth.iloc[i-1]:
            long_smooth.iloc[i] = max(long_stop.iloc[i], long_smooth.iloc[i-1])
        else:
            long_smooth.iloc[i] = long_stop.iloc[i]
        if tr['ha_close'].iloc[i-1] < short_smooth.iloc[i-1]:
            short_smooth.iloc[i] = min(short_stop.iloc[i], short_smooth.iloc[i-1])
        else:
            short_smooth.iloc[i] = short_stop.iloc[i]

    direction = [1]
    for i in range(1, len(tr)):
        if tr['ha_close'].iloc[i] > short_smooth.iloc[i-1]:
            direction.append(1)
        elif tr['ha_close'].iloc[i] < long_smooth.iloc[i-1]:
            direction.append(-1)
        else:
            direction.append(direction[-1])

    tr['dir'] = direction
    tr['prev_dir'] = tr['dir'].shift(1)
    tr['buy_signal'] = (tr['dir'] == 1) & (tr['prev_dir'] == -1)
    tr['sell_signal'] = (tr['dir'] == -1) & (tr['prev_dir'] == 1)
    return tr


def run_backtest(df):
    indicators = calculate_indicators(df)
    trades = []
    balance = START_BALANCE
    position = None
    entry_price = 0

    for i in range(len(indicators)):
        row = indicators.iloc[i]
        time = indicators.index[i]
        price = row['ha_close']
        signal = 'BUY' if row['buy_signal'] else 'SELL' if row['sell_signal'] else None

        if signal:
            if position:
                exit_price = price
                profit = (exit_price - entry_price) * 100 if position == 'BUY' else (entry_price - exit_price) * 100
                balance += profit
                trades.append({
                    'entry_time': entry_time,
                    'exit_time': time,
                    'direction': position,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'profit': profit
                })
            position = signal
            entry_price = price
            entry_time = time

    # Final position close
    if position:
        final_price = indicators['ha_close'].iloc[-1]
        profit = (final_price - entry_price) * 100 if position == 'BUY' else (entry_price - final_price) * 100
        balance += profit
        trades.append({
            'entry_time': entry_time,
            'exit_time': indicators.index[-1],
            'direction': position,
            'entry_price': entry_price,
            'exit_price': final_price,
            'profit': profit
        })

    trades_df = pd.DataFrame(trades)
    return trades_df, balance


def summarize_results(trades, final_balance):
    ticker = SYMBOL
    timeframe = GRANULARITY
    start_time = trades['entry_time'].min()
    end_time = trades['exit_time'].max()
    start_dt = datetime.now().astimezone(LOCAL_TZ)

    # Format trades
    trades_fmt = trades.copy()
    trades_fmt['entry_time'] = trades_fmt['entry_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['exit_time'] = trades_fmt['exit_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['entry_price'] = trades_fmt['entry_price'].round(2)
    trades_fmt['exit_price'] = trades_fmt['exit_price'].round(2)
    trades_fmt['profit'] = trades_fmt['profit'].round(2)

    # Compute equity after each trade
    trades['equity'] = START_BALANCE + trades['profit'].cumsum()

    # Plot equity curve
    plt.figure(figsize=(10, 4))
    plt.plot(trades['exit_time'], trades['equity'], label="Equity", linewidth=2)
    plt.title(f"Equity Curve - {ticker} ({timeframe})")
    plt.xlabel("Time")
    plt.ylabel("Account Equity ($)")
    plt.grid(True)
    plt.tight_layout()
    curve_path = os.path.join(EXPORT_DIR, f"equity_curve_{ticker}_{timeframe}.png")
    plt.savefig(curve_path)
    plt.close()

    # Prepare summary
    summary = {
        'Ticker': ticker,
        'Timeframe': timeframe,
        'Start Time': start_time.strftime("%d/%m/%Y %H:%M:%S"),
        'End Time': end_time.strftime("%d/%m/%Y %H:%M:%S"),
        'Backtest Started': start_dt.strftime("%d/%m/%Y %H:%M:%S"),
        'Backtest Ended': datetime.now().astimezone(LOCAL_TZ).strftime("%d/%m/%Y %H:%M:%S"),
        'Starting Balance': START_BALANCE,
        'Ending Balance': round(final_balance, 2),
        'Total Return': round(final_balance - START_BALANCE, 2),
        'Total Trades': len(trades),
        'Biggest Win': round(trades['profit'].max(), 2),
        'Biggest Loss': round(trades['profit'].min(), 2),
        'Winning Trades': int((trades['profit'] > 0).sum()),
        'Win Rate (%)': round((trades['profit'] > 0).mean(), 4)  # e.g., 0.6458
    }

    print("\n===== BACKTEST SUMMARY =====")
    for key, value in summary.items():
        print(f"{key:<20}: {value}")
    print("============================\n")

    # Save Excel
    os.makedirs(EXPORT_DIR, exist_ok=True)  # Create folder if not exists

    output_path = os.path.join(EXPORT_DIR, f"Chandelier_Backtest_{ticker}_{timeframe}.xlsx")
    wb = Workbook()

    # Trades Sheet
    ws1 = wb.active
    ws1.title = "Trades"
    for r in dataframe_to_rows(trades_fmt, index=False, header=True):
        ws1.append(r)

    # Bold header for Trades sheet
    bold_font = Font(bold=True)
    for cell in ws1[1]:  # First row = header
        cell.font = bold_font

    # Apply formatting to Trades sheet
    for col in ws1.iter_cols(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        header = ws1.cell(row=1, column=col[0].column).value
        for cell in col:
            if header in ['entry_price', 'exit_price']:
                cell.number_format = '0.00'
            elif header == 'profit':
                cell.number_format = '$#,##0.00'

    # Summary Sheet
    ws2 = wb.create_sheet("Summary")
    for k, v in summary.items():
        ws2.append([k, v])

    # Bold header for Summary sheet
    ws2["A1"].font = bold_font
    ws2["B1"].font = bold_font

    # Format Summary sheet values
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=2):
        label = row[0].value
        value_cell = row[1]

        if label in ['Starting Balance', 'Ending Balance', 'Total Return', 'Biggest Win', 'Biggest Loss']:
            value_cell.number_format = '$#,##0.00'
        elif label in ['Total Trades', 'Winning Trades']:
            value_cell.number_format = '0'
        elif label == 'Win Rate (%)':
            value_cell.number_format = '0.00%'

    # Equity Curve Sheet
    ws3 = wb.create_sheet("Equity Curve")
    img = ExcelImage(curve_path)
    img.anchor = "A1"
    ws3.add_image(img)

    wb.save(output_path)
    print(f"\n[OK] Backtest saved to Excel: {output_path}")
    print(f"[OK] Equity curve image saved: {curve_path}")


if __name__ == "__main__":
    print("[*] Fetching OANDA data...")
    df = fetch_oanda_candles()
    print("[*] Running backtest...")
    trades, end_balance = run_backtest(df)
    summarize_results(trades, end_balance)
