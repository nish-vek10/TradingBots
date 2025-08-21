import requests
import pandas as pd
import numpy as np
from datetime import datetime, timezone as dt_timezone
from pytz import timezone
import os
import random
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt

# === USER CONFIG === #
OANDA_TOKEN = "37ee33b35f88e073a08d533849f7a24b-524c89ef15f36cfe532f0918a6aee4c2"
OANDA_ACCOUNT_ID = "101-004-35770497-001"
SYMBOL = "BTC_USD"
GRANULARITY = "M5"
CANDLES = 3840
START_BALANCE = 10000.0
BACKTEST_START_DATE = "2025-05-01T00:00:00Z"
LOCAL_TZ = timezone('Europe/London')
EXPORT_DIR = r"C:\\Users\\anish\\OneDrive\\Desktop\\Anish\\Trading Bots\\EMA-VWAP Scalp"

# === STRATEGY VARIABLES === #
EMA_FAST_PERIOD = 9
EMA_SLOW_PERIOD = 50
RISK_PERCENT = 0.005          # 1.0% risk per trade
SL_MULTIPLIER = 1.5
TP_MULTIPLIER = 2.5

# === FETCH DATA === #
def fetch_oanda_candles():
    max_per_request = 5000
    oanda_url = "https://api-fxpractice.oanda.com/v3/instruments"
    headers = {"Authorization": f"Bearer {OANDA_TOKEN}"}
    all_data = []

    start_time = pd.to_datetime(BACKTEST_START_DATE)
    end_time = datetime.now(dt_timezone.utc).replace(microsecond=0)

    prev_last_time = None

    while start_time < end_time:
        params = {
            "granularity": GRANULARITY,
            "price": "M",
            "from": start_time.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "count": max_per_request
        }

        url = f"{oanda_url}/{SYMBOL}/candles"
        print(f"[*] Fetching candles starting {start_time}")

        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            raise RuntimeError(f"Failed to fetch candles: {response.status_code} - {response.text}")

        candles = response.json().get("candles", [])
        if not candles:
            break

        for c in candles:
            if c['complete']:
                utc_time = pd.to_datetime(c['time']).tz_convert(LOCAL_TZ)
                all_data.append({
                    "time": utc_time,
                    "open": float(c['mid']['o']),
                    "high": float(c['mid']['h']),
                    "low": float(c['mid']['l']),
                    "close": float(c['mid']['c']),
                    "volume": int(c['volume'])
                })

        last_time = pd.to_datetime(candles[-1]['time'])

        if prev_last_time is not None and last_time <= prev_last_time:
            print("[!] No new data returned. Stopping fetch loop.")
            break

        prev_last_time = last_time
        start_time = last_time + pd.Timedelta(seconds=1)


    if not all_data:
        print("[!] No candle data fetched.")
        return pd.DataFrame()

    df = pd.DataFrame(all_data)
    df.set_index('time', inplace=True)
    return df

df = fetch_oanda_candles()
if df.empty:
    raise RuntimeError("No data fetched from OANDA. Backtest aborted.")

# === CALCULATE INDICATORS === #
def calculate_indicators(df):
    df['ema_fast'] = df['close'].ewm(span=EMA_FAST_PERIOD, adjust=False).mean()
    df['ema_slow'] = df['close'].ewm(span=EMA_SLOW_PERIOD, adjust=False).mean()

    df['vwap'] = (df['volume'] * (df['high']+df['low']+df['close'])/3).cumsum() / df['volume'].cumsum()
    df['tr'] = np.maximum(df['high']-df['low'],
                          np.maximum(abs(df['high']-df['close'].shift(1)),
                                     abs(df['low']-df['close'].shift(1))))
    df['atr'] = df['tr'].rolling(window=14).mean()
    return df

# === BACKTEST STRATEGY === #
def run_backtest(df):
    trades = []
    balance = START_BALANCE
    position = None
    entry_price = 0
    sl = 0
    tp = 0
    lot_size = 1.0

    for i in range(51, len(df)):
        price = df['close'].iloc[i]
        row = df.iloc[i]

        # ENTRY CONDITIONS
        if position is None:
            risk_per_trade = balance * RISK_PERCENT

            # LONG SETUP
            if (row['close'] > row['vwap'] and
                row['ema_fast'] > row['ema_slow'] and
                row['close'] > row['ema_fast'] and
                df['low'].iloc[i-1] <= row['ema_fast']):

                sl = price - (SL_MULTIPLIER * row['atr'])
                risk_pips = price - sl
                lot_size = risk_per_trade / risk_pips if risk_pips > 0 else 1.0
                tp = price + (TP_MULTIPLIER * risk_pips)
                position = 'BUY'
                entry_price = price
                entry_time = df.index[i]

            # SHORT SETUP
            elif (row['close'] < row['vwap'] and
                  row['ema_fast'] < row['ema_slow'] and
                  row['close'] < row['ema_fast'] and
                  df['high'].iloc[i-1] >= row['ema_fast']):

                sl = price + (SL_MULTIPLIER * row['atr'])
                risk_pips = sl - price
                lot_size = risk_per_trade / risk_pips if risk_pips > 0 else 1.0
                tp = price - (TP_MULTIPLIER * risk_pips)
                position = 'SELL'
                entry_price = price
                entry_time = df.index[i]

        # EXIT CONDITIONS
        else:
            if position == 'BUY' and (row['low'] <= sl or row['high'] >= tp):
                exit_price = sl if row['low'] <= sl else tp
                profit = (exit_price - entry_price) * lot_size
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': 'BUY', 'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

            elif position == 'SELL' and (row['high'] >= sl or row['low'] <= tp):
                exit_price = sl if row['high'] >= sl else tp
                profit = (entry_price - exit_price) * lot_size
                balance += profit
                trades.append({'entry_time': entry_time, 'exit_time': df.index[i], 'direction': 'SELL', 'entry_price': entry_price, 'exit_price': exit_price, 'profit': profit})
                position = None

    trades_df = pd.DataFrame(trades)
    return trades_df, balance

# === SUMMARIZE RESULTS === #
def summarize_results(trades, final_balance):
    ticker = SYMBOL
    timeframe = GRANULARITY
    start_time = trades['entry_time'].min()
    end_time = trades['exit_time'].max()

    # Format trade data
    trades_fmt = trades.copy()
    trades_fmt['entry_time'] = trades_fmt['entry_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['exit_time'] = trades_fmt['exit_time'].dt.strftime("%d/%m/%Y %H:%M:%S")
    trades_fmt['entry_price'] = trades_fmt['entry_price'].round(2)
    trades_fmt['exit_price'] = trades_fmt['exit_price'].round(2)
    trades_fmt['profit'] = trades_fmt['profit'].round(2)

    # Equity curve
    trades['equity'] = START_BALANCE + trades['profit'].cumsum()
    plt.figure(figsize=(10, 4))
    plt.plot(trades['exit_time'], trades['equity'], label="Equity", linewidth=2)
    plt.title(f"Equity Curve - {ticker} ({timeframe})")
    plt.xlabel("Time")
    plt.ylabel("Account Equity ($)")
    plt.grid(True)
    plt.tight_layout()
    curve_path = os.path.join(EXPORT_DIR,
                              f"equity_curve_{ticker}_{timeframe}_EMA({EMA_FAST_PERIOD}-{EMA_SLOW_PERIOD}).png")
    plt.savefig(curve_path)
    plt.close()

    # Calculate performance stats
    percentage_return = ((final_balance - START_BALANCE) / START_BALANCE) * 100
    gross_profit = trades.loc[trades['profit'] > 0, 'profit'].sum()
    gross_loss = trades.loc[trades['profit'] < 0, 'profit'].sum()
    profit_factor = gross_profit / abs(gross_loss) if gross_loss != 0 else float('inf')

    summary = {
        'Ticker': ticker,
        'Timeframe': timeframe,
        'Start Time': start_time.strftime("%d/%m/%Y %H:%M:%S"),
        'End Time': end_time.strftime("%d/%m/%Y %H:%M:%S"),
        'Starting Balance': f"${START_BALANCE:,.2f}",
        'Ending Balance': f"${final_balance:,.2f}",
        'Total Return': f"${(final_balance - START_BALANCE):,.2f}",
        'Return (%)': f"{percentage_return:.2f}%",
        'Total Trades': int(len(trades)),
        'Biggest Win': f"${trades['profit'].max():,.2f}",
        'Biggest Loss': f"${trades['profit'].min():,.2f}",
        'Winning Trades': int((trades['profit'] > 0).sum()),
        'Win Rate (%)': f"{(trades['profit'] > 0).mean() * 100:.2f}%",
        'Profit Factor': f"{profit_factor:.2f}",
        '\n--- STRATEGY PARAMETERS ---': '',
        'EMA Fast Period': EMA_FAST_PERIOD,
        'EMA Slow Period': EMA_SLOW_PERIOD,
        'SL Multiplier': SL_MULTIPLIER,
        'TP Multiplier': TP_MULTIPLIER,
        'Risk per Trade (%)': f"{RISK_PERCENT * 100:.2f}%"
    }

    print("\n===== BACKTEST SUMMARY =====")
    for key, value in summary.items():
        print(f"{key:<20}: {value}")
    print("============================\n")

    os.makedirs(EXPORT_DIR, exist_ok=True)
    output_path = os.path.join(EXPORT_DIR,
                               f"EMA_VWAP_Scalping_{ticker}_{timeframe}_EMA({EMA_FAST_PERIOD}-{EMA_SLOW_PERIOD}-{random.randint(1,20)}).xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Trades"
    for r in dataframe_to_rows(trades_fmt, index=False, header=True):
        ws1.append(r)

    bold_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = bold_font

    ws2 = wb.create_sheet("Summary")
    for k, v in summary.items():
        ws2.append([k, v])

    ws3 = wb.create_sheet("Equity Curve")
    img = ExcelImage(curve_path)
    img.anchor = "A1"
    ws3.add_image(img)

    wb.save(output_path)
    print(f"\n[OK] Backtest saved to Excel: {output_path}")
    print(f"[OK] Equity curve image saved: {curve_path}")

if __name__ == "__main__":
    print("[*] Fetching OANDA data...")
    df = fetch_oanda_candles()

    # Safety check to avoid 'NoneType' errors
    if df.empty:
        raise RuntimeError("No data fetched from OANDA. Backtest aborted.")

    # Only calculate indicators if data exists
    df = calculate_indicators(df)

    print("[*] Running backtest...")
    trades, end_balance = run_backtest(df)
    summarize_results(trades, end_balance)
